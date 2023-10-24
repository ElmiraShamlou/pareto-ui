from zipfile import ZipFile
import pprint
import xml.sax, xml.sax.handler
import shutil
import math
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# pandas: geoparse
def ParseKMZ(filename = "Demo_network_correct.kmz"):
    global NODE_NAMES
    NODE_NAMES = set()

    def calculate_distance(coord1, coord2):
        # print(f'calculating distance from {coord1} to {coord2}')
        distance = math.sqrt(((float(coord1[0]) - float(coord2[0]))**2) + ((float(coord1[1]) - float(coord2[1]))**2))
        return distance

    ## check if kmz or kml
    if filename[-3:] == 'kmz':
        kmz = ZipFile(filename, 'r')
        kml = kmz.open('doc.kml', 'r')
    elif filename[-3:] == 'kml':
        print('file type is kml')
        kml = open(filename, 'r')
    else:
        print('filetype is not kmz or kml')
        return {}

    class PlacemarkHandler(xml.sax.handler.ContentHandler):
        def __init__(self):
            self.inName = False # handle XML parser events
            self.inPlacemark = False
            self.mapping = {}
            self.buffer = ""
            self.name_tag = ""
        
        def startElement(self, name, attributes):
            if name == "Placemark": # on start Placemark tag
                self.inPlacemark = True
                self.buffer = ""
            if self.inPlacemark:
                if name == "name": # on start title tag
                    self.inName = True # save name text to follow
            
        def characters(self, data):
            if self.inPlacemark: # on text within tag
                self.buffer += data # save text if in title
            
        def endElement(self, name):
            self.buffer = self.buffer.strip('\n\t')
        
            if name == "Placemark":
                self.inPlacemark = False
                self.name_tag = "" #clear current name
        
            elif name == "name" and self.inPlacemark:
                self.inName = False # on end title tag           
                self.name_tag = self.buffer.strip()
                while self.name_tag in NODE_NAMES:
                    self.name_tag = f'{self.name_tag}_'
                NODE_NAMES.add(self.name_tag)
                self.mapping[self.name_tag] = {}
            elif self.inPlacemark:
                if name in self.mapping[self.name_tag]:
                    self.mapping[self.name_tag][name] += self.buffer
                else:
                    self.mapping[self.name_tag][name] = self.buffer
            self.buffer = ""

    parser = xml.sax.make_parser()
    handler = PlacemarkHandler()
    parser.setContentHandler(handler)
    parser.parse(kml)
    kmz.close()

    # this is the object that contains all the data for each point on the map
    result_object = handler.mapping

    data = {}
    all_nodes = {}
    production_pads = {}
    completion_pads = {}
    network_nodes = {}
    disposal_sites = {}
    treatment_sites = {}
    storage_sites = {}
    freshwater_sources = {}
    other_nodes = {}
    arcs = {}

    for key in result_object:
        # data[key] = {}
        coordinates_string = result_object[key]["coordinates"]
        coordinates_split = coordinates_string.split(' ')
        coordinates_list = []
        for each in coordinates_split:
            # coords = []
            if len(each) > 0:
                coordinates_list.append(each.split(','))
        result_object[key]["coordinates"] = coordinates_list

        if len(coordinates_list) > 1:
            result_object[key]["node_type"] = "path"
            arcs[key] = result_object[key]

        else:
            result_object[key]["node_type"] = "point"
            result_object[key]["coordinates"] = result_object[key]["coordinates"][0]
            ## determine what kind of node it is:
            if key[0:2].upper() == 'PP':
                production_pads[key] = result_object[key]
            elif key[0:2].upper() == 'CP':
                completion_pads[key] = result_object[key]
            elif key[0].upper() == 'N':
                network_nodes[key] = result_object[key]
            elif key[0].upper() == 'K':
                disposal_sites[key] = result_object[key]
            elif key[0].upper() == 'R':
                treatment_sites[key] = result_object[key]
            elif key[0].upper() == 'S' and len(key) < 4:
                storage_sites[key] = result_object[key]
            elif key[0].upper() == 'F':
                freshwater_sources[key] = result_object[key]
            else:
                other_nodes[key] = result_object[key]
            all_nodes[key] = result_object[key]

    connections = {
        "all_connections_list": [],
        "all_connections": {},
        "P": {
            "C": [],
            "N": [],
            "K": [],
        }, 
        "C": {
            "C": [],
            "N": [],
            "K": [],
            "S": [],
        }, 
        "N": {
            "C": [],
            "N": [],
            "K": [],
            "S": [],
            "R": [],
            "O": [],
            "P": [],
        },
        "S": {
            "C": [],
            "N": [],
            "O": [],
        }, 
        "R": {
            "C": [],
            "N": [],
            "O": [],
            "S": [],
            "R": [],
        }, 
        "F": {
            "C": [],
        }
    }

    ## possible connection types:
    # P - N, C, K
    # C - N, C, K, S
    # N - N, C, K, R, S, O
    # S - N, O, C
    # R - C, S, N, O
    # F - C
    # K - 
    ## cannot determine if trucking or piped; ASSUME ALL ARE PIPED for now


    ## for each arc endpoint, determine the nearest node
    for arc_key in arcs:
        arc = arcs[arc_key]
        node_list = []
        for arc_coordinates in arc["coordinates"]:
            min_distance = 100000.0
            closest_node = ""
            # check each node
            for node_key in all_nodes:
                node = all_nodes[node_key]
                node_coordinates = node["coordinates"]
                distance = calculate_distance(arc_coordinates, node_coordinates)
                if distance < min_distance:
                    closest_node = node_key
                    min_distance = distance
                    # print(f'closest node is {node_key}')
            if len(node_list) > 0:
                ## add connection
                origin_node = node_list[-1]
                connection = [origin_node, closest_node]
                # origin_node_data = all_nodes[origin_node]
                # destination_node_data = all_nodes[closest_node]

                connections["all_connections_list"].append(connection)

                ## ASSUME connections are bidirectional
                if origin_node in connections["all_connections"]:
                    connections["all_connections"][origin_node].append(closest_node)
                else:
                    connections["all_connections"][origin_node] = [closest_node]
                if closest_node in connections["all_connections"]:
                    connections["all_connections"][closest_node].append(origin_node)
                else:
                    connections["all_connections"][closest_node] = [origin_node]


                try:
                    connections[origin_node[0]][closest_node[0]].append(connection)
                except:
                    print(f'unable to add connection: {connection}')



            node_list.append(closest_node)
        arc['node_list'] = node_list

    data['all_nodes'] = all_nodes
    data['ProductionPads'] = production_pads
    data['CompletionsPads'] = completion_pads
    data['NetworkNodes'] = network_nodes
    data['SWDSites'] = disposal_sites
    data['TreatmentSites'] = treatment_sites
    data['StorageSites'] = storage_sites
    data['FreshwaterSources'] = freshwater_sources
    data['other_nodes'] = other_nodes
    data['arcs'] = arcs
    data['connections'] = connections


    return data 


def WriteKMZDataToExcel(data, output_file_name="kmz_scenario"):
    input_path = "pareto_input_template.xlsx"
    excel_path = f'{output_file_name}.xlsx'
    print(f'writing data to excel at {excel_path}')

    ## step 1: copy pareto_input_template to new file for writing
    shutil.copyfile(input_path, excel_path)

    ## step 2: open excel workbook
    wb = load_workbook(excel_path, data_only=True)

    ## step 3: add nodes
    node_keys = [
        'ProductionPads', 'CompletionsPads', 'SWDSites', 'FreshwaterSources', 
        'StorageSites', 'TreatmentSites', 'NetworkNodes'
    ]

    column = 1
    for node_key in node_keys:
        row = 2
        ws = wb[node_key]
        for node in data[node_key]:
            print(f'{node_key}: adding {node}')
            cellLocation = f'{get_column_letter(column)}{row}'
            ws[cellLocation] = node
            row+=1

    ## step 4: add arcs
    piped_arcs = {
        "PNA": ["ProductionPads", "NetworkNodes"],
        "CNA": ["CompletionsPads", "NetworkNodes"],
        "CCA": ["CompletionsPads", "CompletionsPads"],
        "NNA": ["NetworkNodes", "NetworkNodes"],
        "NCA": ["NetworkNodes", "CompletionsPads"],
        "NKA": ["NetworkNodes", "SWDSites"],
        "NRA": ["NetworkNodes", "TreatmentSites"],
        "NSA": ["NetworkNodes", "StorageSites"],
        # "NOA": ["NetworkNodes", "BeneficialReuse?"],
        "SNA": ["StorageSites", "NetworkNodes"],
        # "SOA": ["StorageSites", "BeneficialReuse"],
        "FCA": ["FreshwaterSources", "CompletionsPads"],
        "RCA": ["TreatmentSites", "CompletionsPads"],
        "RSA": ["TreatmentSites", "StorageSites"],
        "SCA": ["StorageSites", "CompletionsPads"],
        "RNA": ["TreatmentSites", "NetworkNodes"],
        # "ROA": ["TreatmentSites", "BeneficialReuse"],
    }
    piped_arc = "PNA" ## will be a loop
    for piped_arc in piped_arcs:
        ws = wb[piped_arc]
        piped_arc_node1 = piped_arcs[piped_arc][0]
        piped_arc_node2 = piped_arcs[piped_arc][1]
        column = 1
        row = 3
        row_nodes = []
        print(f'{piped_arc}: adding {piped_arc_node1}')
        for node in data[piped_arc_node1]:
            cellLocation = f'{get_column_letter(column)}{row}'
            ws[cellLocation] = node
            row_nodes.append(node)
            row+=1
        column = 2
        row = 2
        print(f'{piped_arc}: adding {piped_arc_node2}')
        for node in data[piped_arc_node2]:
            cellLocation = f'{get_column_letter(column)}{row}'
            ws[cellLocation] = node
            # print('checking for connections')
            ind = 3
            for row_node in row_nodes:
                if row_node in data["connections"]["all_connections"]:
                    if node in data["connections"]["all_connections"][row_node]:
                        # print(f'adding connection for {row_node}:{node}')
                        cellLocation = f'{get_column_letter(column)}{ind}'
                        # print(f'adding to cell location: {cellLocation}')
                        ws[cellLocation] = 1

                ind+=1
            column+=1


    ## step 5: Save and close
    wb.save(excel_path)
    wb.close()


# data = ParseKMZ(filename="Demo_network_correct.kmz")
# print('got data')
# pp = pprint.PrettyPrinter(indent=1)
# pp.pprint(data['connections']['all_connections'])
# WriteKMZDataToExcel(data)
